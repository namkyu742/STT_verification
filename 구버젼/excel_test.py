from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side, Alignment

result_string = [
        "----------------------- | RESULT | -----------------------\n",
        "success count : \n",
        "success ratio : \n",
        "Runtime       : \n"
    ]
failed_list = []
failed_list.append({'name':"str(temp_name)", 'ratio':"str(ratio)"})








write_wb = Workbook()
write_ws = write_wb.active

write_ws.merge_cells('A1:C1')
write_ws.cell(1, 1, "정확도 95% 이하 음성데이터 목록")
write_ws.cell(2, 1, "No")
write_ws.cell(2, 2, "음성데이터")
write_ws.cell(2, 3, "정확도")

write_ws.column_dimensions['A'].width = 10
write_ws.column_dimensions['B'].width = 20
write_ws.column_dimensions['C'].width = 20

count = 3
for i in failed_list:
    write_ws.cell(count, 1, count-2)
    write_ws.cell(count, 2, i['name'])
    write_ws.cell(count, 3, i['ratio'])
    count += 1

# write_ws.merge_cells(count+0, 1, count+0, 3)
# write_ws.merge_cells(count+1, 1, count+1, 3)
# write_ws.merge_cells(count+2, 1, count+2, 3)
# write_ws.merge_cells(count+3, 1, count+3, 3)
for i in range(0, 4):
    write_ws.merge_cells(f'A{count+i}:C{count+i}')
    write_ws.cell(count+i, 1, result_string[i])


THIN_BORDER = Border(Side('thin'),Side('thin'),Side('thin'),Side('thin'))
for rng in write_ws[f'A1:C{count+3}']:
    for cell in rng:
        cell.border = THIN_BORDER

write_wb.save('./test.xlsx')