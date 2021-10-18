#-*- coding:utf-8 -*-
import urllib3
import json
import base64
import json
import os


def use_stt_api(file_name):
    openApiURL = "http://aiopen.etri.re.kr:8000/WiseASR/Recognition"
    accessKey = "198b2f86-c3a3-409c-b524-3f065eb25dd7"
    languageCode = "korean"

    audioFilePath = os.path.join(os.path.dirname(__file__), 'aihub_data', 'hobby_01', '001', file_name + '.wav')
    
    file = open(audioFilePath, "rb")
    audioContents = base64.b64encode(file.read()).decode("utf8")
    file.close()
    
    requestJson = {
        "access_key": accessKey,
        "argument": {
            "language_code": languageCode,
            "audio": audioContents
        }
    }
    
    http = urllib3.PoolManager()
    response = http.request(
        "POST",
        openApiURL,
        headers={"Content-Type": "application/json; charset=UTF-8"},
        body=json.dumps(requestJson)
    )

    data = json.loads(response.data.decode("utf-8", errors='ignore'))    

    # [END speech_quickstart]
    return data['return_object']['recognized']


def check_gantueo(text):
    # [간투어 표현 : 발성자가 다음 발성을 준비하기 위해서 소요되는 시간을 벌기 위해서 발성하는 것]
    # [간투어 체크 시작]
    import re
    text = re.sub("아니/", "아니", text)
    text = re.sub("저기/", "저기", text)
    text = re.sub("아/", "아", text)
    text = re.sub("그/", "그", text)
    text = re.sub("어/", "어", text)
    text = re.sub("음/", "음", text)
    text = re.sub("저/", "저", text)
    text = re.sub("으/", "으", text)
    text = re.sub("웅/", "웅", text)
    text = re.sub("에/", "에", text)
    # [간투어 체크 종료]
    return text

def check_Double_Copying(text1, text2):
    # 이중전사 검사
    # 이중전사 : (철자전사)/(발음전사)
    import regex
    from difflib import SequenceMatcher

    list1 = regex.findall('\p{Hangul}+|\W|\d', text1)
    flag1 = False
    flag2 = False

    text_result = ""
    temp_text1 = ""
    temp_text2 = ""
        
    for i in range(0, len(list1)):
        if (list1[i] == '(' and flag1 == False and flag2 == False):
            flag1 = True
            continue
        if (list1[i] == ')' and flag1 == True and flag2 == False):
            flag2 = False
            continue
        if (list1[i] == '/' and flag1 == False and flag2 == False):
            continue
        if (list1[i] == '/' and flag1 == True and flag2 == False):
            flag1 = False
            flag2 = True
            continue
        if (list1[i] == '(' and flag1 == False and flag2 == True):
            flag1 = True
            continue
        if (list1[i] == ')' and flag1 == True and flag2 == True):
            flag1 = False
            flag2 = False
            # 비교문에 해당 문자열이 존재하는지 검사
            result1 = SequenceMatcher(None, temp_text1, text2).ratio()
            result2 = SequenceMatcher(None, temp_text2, text2).ratio()

            # 검사 결과에 따라 해당 문자열을 추가
            if (result1 > result2):
                text_result += temp_text1
            elif (result1 < result2):
                text_result += temp_text2
            else:
                # 원문과 유사도 검사 결과가 둘이 동일한 경우, 문자열1을 채택
                text_result += temp_text1 

            temp_text1 = ""
            temp_text2 = ""
            continue

        if (flag1 == True and flag2 == False):
            # 이중전사 1
            temp_text1 += list1[i]
        elif (flag1 == True and flag2 == True):
            # 이중전사 2
            temp_text2 += list1[i]

        # 그 외의 글자들은 그대로 복사
        if (flag1 == False):
            text_result += list1[i]        

    return text_result

def check_Spacing_and_PunctuationMarks(text):
    # 띄어쓰기 제거
    # 문장부호 제거
    #   쉼표 ,
    #   느낌표 !
    #   물음표 ?
    #   마침표 .

    # 중복발성 기호 제거 '+'
    # 잡음 기호 제거 '*'

    text_result = ""
    for i in range(0, len(text)):
        if (text[i] == ' ' 
        or text[i] == '+' 
        or text[i] == '*'
        or text[i] == ',' 
        or text[i] == '!' 
        or text[i] == '?' 
        or text[i] == '.' ):
            continue        
        text_result = text_result + text[i]

    return text_result

def text_difflib(file_name_origin, transScript):
    from difflib import SequenceMatcher
    import os
    
    file_name = os.path.join(
        os.path.dirname(__file__), 'aihub_data', 'hobby_01', '001', file_name_origin + '.txt')

    f1 = open(file_name, mode='rt', encoding='UTF8')

    text1 = f1.read()

    text2 = transScript
    
    # 영어 소문자를 대문자로 변경
    text2 = text2.upper()

    text3 = check_Double_Copying(text1, text2)
    
    text3 = check_Spacing_and_PunctuationMarks(text3)
    text4 = check_Spacing_and_PunctuationMarks(text2)
    ratio = SequenceMatcher(None, text3, text4).ratio()
    
    if (ratio < 0.95):
        print('[RESULT] :', file_name_origin)
        print('OriginalScript    :', text1)
        print('TranScript        :', text2)
        print('OriginalScript[T] :', text3)
        print('TranScript[T]     :', text4)
        print('RATIO :', ratio)
        print('----------------------------------------------------------')

    return ratio

def one_click(file_name):
    transScript = use_stt_api(file_name)
    return text_difflib(file_name, transScript)

def ShowRunTime(runningTime):
    run_second = int(runningTime)
    run_minute = 0
    run_hour = 0
    if (run_second > 3600):
        run_hour = int(run_second / 3600)
        run_second = run_second - (3600*run_hour)
    if (run_second > 60):
        run_minute = int(run_second / 60)
        run_second = run_second - (60*run_minute)

    run_time = ""
    if (run_hour > 0):
        run_time += str(run_hour) + "시간 "
    if (run_minute > 0):
        run_time += str(run_minute) + "분 "
    run_time += str(run_second) + "초"

    return run_time

def Output_result_as_EXCEL(failed_list, result_string):

    # failed_fp = open("failed_list.dat", "wt")
    # for i in failed_list:
    #     failed_fp.write(str(i))
    #     failed_fp.write('\n')
    #     print(i['name'])
    #     print(i['ratio'])

    # failed_fp.write(result_string)
    # failed_fp.close()

    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, Font, Alignment
    write_wb = Workbook()
    write_ws = write_wb.active

    write_ws.merge_cells('A1:C1')
    write_ws.cell(1, 1, "정확도 95% 이하 음성데이터 목록").font = Font(size=15, bold=True)
    write_ws.cell(1, 1).alignment = Alignment(horizontal="center")
    write_ws.cell(2, 1, "No").alignment = Alignment(horizontal="center")
    write_ws.cell(2, 2, "음성데이터").alignment = Alignment(horizontal="center")
    write_ws.cell(2, 3, "정확도").alignment = Alignment(horizontal="center")
    
    write_ws.column_dimensions['A'].width = 10
    write_ws.column_dimensions['B'].width = 20
    write_ws.column_dimensions['C'].width = 20

    count = 3
    for i in failed_list:
        write_ws.cell(count, 1, count-2)
        write_ws.cell(count, 2, i['name'])
        write_ws.cell(count, 3, i['ratio'])
        count += 1

    for i in range(0, 4):
        write_ws.merge_cells(f'A{count+i}:C{count+i}')
        write_ws.cell(count+i, 1, result_string[i])


    THIN_BORDER = Border(Side('thin'),Side('thin'),Side('thin'),Side('thin'))
    for rng in write_ws[f'A1:C{count+3}']:
        for cell in rng:
            cell.border = THIN_BORDER

    write_wb.save('./test.xlsx')









if __name__ == '__main__':
    # [For check Runtime]
    import time
    start = time.time()

    failed_list = []
    count = 0
    file_name = "hobby_0000"
    start_num = 2
    end_num = 2
    num_size = end_num - start_num + 1
    if (num_size<=0):
        # [Error : Input range]
        exit()

    for i in range(start_num, end_num + 1):
        # [For Adjustment file name]
        if i<10:
            temp_name = file_name + '000' + str(i)
        elif 10<=i<100:
            temp_name = file_name + '00' + str(i)
        elif 100<=i<1000:
            temp_name = file_name + '0' + str(i)




        ratio = one_click(temp_name)
        if ratio>=0.95:
            count += 1
        else:
            failed_list.append({'name':str(temp_name), 'ratio':str(ratio)})

    result_string = "----------------------- | RESULT | -----------------------\n"
    result_string += "success count : " + str(count) + " / " + str(num_size) + "\n"
    result_string += "success ratio : " + str(round(count/num_size*100)) + "%\n"
    result_string += "Runtime       :" + ShowRunTime(round(time.time() - start, 0))
    print(result_string)

    result_string2 = [
        "----------------------- | RESULT | -----------------------\n",
        "success count : " + str(count) + " / " + str(num_size) + "\n",
        "success ratio : " + str(round(count/num_size*100)) + "%\n",
        "Runtime       :" + ShowRunTime(round(time.time() - start, 0))
    ]


    Output_result_as_EXCEL(failed_list, result_string2)


