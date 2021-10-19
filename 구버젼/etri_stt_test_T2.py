#-*- coding:utf-8 -*-

# 프로젝트 시작 2021년도 9월
# ------------------------------------------------------------------
# [ToDo 목록]
# 1. 완료 - 함수 클래스화
# 2. 완료 - 20초 이상의 음성데이터 분할처리
# 3. 완료 - 결과를 액셀파일로 출력
# 4. ffmpeg 경고문제 해결하기
# 5. 액셀파일이 열려있을때 대처하기
# 6. 액셀파일에 원문과 전사문 다른점 강조표시
# 7. 쓰레드 사용해서 작업 분담
# 8. 
# 
# ------------------------------------------------------------------
# [작성자]
#   조남규 (namkyu742@naver.com)

# ------------------------------------------------------------------
# [참고자료]
# 예외처리 (https://dogpitch.oopy.io/python/file-io-and-exception)
# 음성데이터 분할 (https://stackoverflow.com/questions/37999150/how-to-split-a-wav-file-into-multiple-wav-files)
# os 라이브러리 사용법 (https://webisfree.com/2018-03-16/python-%ED%8C%8C%EC%9D%BC-%EB%B0%8F-%EB%94%94%EB%A0%89%ED%86%A0%EB%A6%AC-%EC%82%AD%EC%A0%9C%ED%95%98%EB%8A%94-%EB%B0%A9%EB%B2%95)
# 실행파일 만들기1 (https://wikidocs.net/21952)
# 실행파일 만들기2 (https://m.blog.naver.com/smilewhj/221070338758)


from numpy import e
import urllib3
import json
import base64
import json
import os
from pydub import AudioSegment
import math
import threading
import time

# [CONSTANT VARIALBE]
PRINT_RESULT_OFF = 0
PRINT_RESULT_ON = 1

NUM_OF_THREADS = 1
TIME_OF_SLEEP = 1.0

count = 0
failed_list = []
lock = threading.Lock()

# ------------------------------------------------------------------

class ETRI_STT_API():
    def __init__(self, folderpath, filename):
        self.folderpath = folderpath
        self.filename = filename
        self.transScript = ""
        self.ratio = 0.0
        
    def get_trans_script(self):
        return self.transScript

    def set_trans_script(self, trans_script):
        self.transScript = trans_script

    def get_ratio(self):
        return self.ratio
        
    def get_filepath(self):
        return os.path.join(self.folderpath, self.filename + ".wav")

    def get_origin_script(self):
        filepath = os.path.join(self.folderpath, self.filename + ".txt")
        script_file = open(filepath, mode='rt', encoding='UTF8')
        return script_file.read()

    def use_STT_API(self):
        openApiURL = "http://aiopen.etri.re.kr:8000/WiseASR/Recognition"
        accessKey = "198b2f86-c3a3-409c-b524-3f065eb25dd7"
        languageCode = "korean"

        audioFilePath = self.get_filepath()
        try:
            file = open(audioFilePath, "rb")
        except Exception as e:
            print("Exception:", e)
            exit()

        audioContents = base64.b64encode(file.read()).decode("utf8")
        file.close()
        
        requestJson = {
            "access_key": accessKey,
            "argument": {
                "language_code": languageCode,
                "audio": audioContents
            }
        }
        time.sleep(TIME_OF_SLEEP)
        http = urllib3.PoolManager()
        response = http.request(
            "POST",
            openApiURL,
            headers={"Content-Type": "application/json; charset=UTF-8"},
            body=json.dumps(requestJson)
        )

        data = json.loads(response.data.decode("utf-8", errors='ignore'))    
        print("----", data)
        return data['return_object']['recognized']

    def check_Double_Copying(self):
        # 이중전사 검사
        # 이중전사 : (철자전사)/(발음전사)
        import regex
        from difflib import SequenceMatcher

        trans_script = self.transScript
        list1 = regex.findall('\p{Hangul}+|\W|\d', self.get_origin_script())
        # flag 변수 초기화
        flag1 = False
        flag2 = False

        result_script = ""
        temp_script_1 = ""
        temp_script_2 = ""
            
        for i in range(0, len(list1)):
            if (list1[i] == '(' and flag1 == False and flag2 == False):
                flag1 = True
                continue
            if (list1[i] == ')' and flag1 == True and flag2 == False):
                flag2 = False
                continue
            if (list1[i] == '/' and flag1 == False and flag2 == False):
                continue
            if (list1[i] == '/' and flag1 == True and flag2 == False):
                flag1 = False
                flag2 = True
                continue
            if (list1[i] == '(' and flag1 == False and flag2 == True):
                flag1 = True
                continue
            if (list1[i] == ')' and flag1 == True and flag2 == True):
                flag1 = False
                flag2 = False
                # 전사문에 해당 문자열이 존재하는지 검사
                result1 = SequenceMatcher(None, temp_script_1, trans_script).ratio()
                result2 = SequenceMatcher(None, temp_script_2, trans_script).ratio()

                # 검사 결과에 따라 해당 문자열을 추가
                # 원문과 유사도 검사 결과가 둘이 동일한 경우, 문자열1을 채택
                if (result1 > result2): result_script += temp_script_1
                elif (result1 < result2): result_script += temp_script_2
                else: result_script += temp_script_1 

                # temp_script 초기화
                temp_script_1 = ""
                temp_script_2 = ""
                continue

            if (flag1 == True and flag2 == False):  # temp_script_1에 철자전사 기록
                temp_script_1 += list1[i]
            elif (flag1 == True and flag2 == True): # temp_script_2에 발음전사 기록
                temp_script_2 += list1[i]
            else:       # 이중전사에 해당되지 않는 글자는 그대로 result_script에 기록
                result_script += list1[i]        

        return result_script

    def check_Spacing_and_PunctuationMarks(self, script):
        # 다음의 기호들을 제거
        # 공백
        # 쉼표                  ,
        # 느낌표                !
        # 물음표                ?
        # 마침표                .
        # 중복발성 기호 제거     +
        # 잡음 기호 제거         *

        result_script = ""
        for i in range(0, len(script)):
            if (script[i] == ' ' 
            or script[i] == '+' 
            or script[i] == '*'
            or script[i] == ',' 
            or script[i] == '!' 
            or script[i] == '?' 
            or script[i] == '.' ):
                continue        
            result_script = result_script + script[i]
        return result_script

    def analyze_script_difference(self, option):
        from difflib import SequenceMatcher
    
        text1 = self.get_origin_script()
        text2 = self.transScript
        
        text2 = text2.upper()   # 영어 소문자를 대문자로 변경
        text3 = self.check_Double_Copying()
        
        text3 = self.check_Spacing_and_PunctuationMarks(text3)
        text4 = self.check_Spacing_and_PunctuationMarks(text2)
        
        ratio = SequenceMatcher(None, text3, text4).ratio()
        # if (option == 1):
        #     if (ratio < 0.95):
        print('[RESULT] :', self.filename)
        print('OriginalScript    :', text1)
        print('TranScript        :', text2)
        print('OriginalScript[T] :', text3)
        print('TranScript[T]     :', text4)
        print('RATIO :', ratio)
        print('------------------------------------------------------------------')

        return ratio

    def one_click(self):
        self.transScript = self.use_STT_API()
        self.ratio = self.analyze_script_difference(PRINT_RESULT_ON)

    def one_click_splited(self):
        self.transScript = self.use_STT_API()
        
def ShowRunTime(runningTime):
    run_second = int(runningTime)
    run_minute = 0
    run_hour = 0
    if (run_second > 3600):
        run_hour = int(run_second / 3600)
        run_second = run_second - (3600*run_hour)
    if (run_second > 60):
        run_minute = int(run_second / 60)
        run_second = run_second - (60*run_minute)

    run_time = ""
    if (run_hour > 0):
        run_time += str(run_hour) + "시간 "
    if (run_minute > 0):
        run_time += str(run_minute) + "분 "
    run_time += str(run_second) + "초"

    return run_time

def Output_result_as_EXCEL(failed_list, result_list):
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
    write_wb = Workbook()
    write_ws = write_wb.active

    # [Title]
    write_ws.merge_cells('A1:O1')
    write_ws.cell(1, 1, "정확도 95% 이하 음성데이터 목록").alignment = Alignment(horizontal="center", vertical="center")
    write_ws.cell(1, 1).font = Font(size=15, bold=True)
    write_ws.row_dimensions[1].height = 40

    # [액셀 열 간격 조정]
    write_ws.column_dimensions['A'].width = 6
    write_ws.column_dimensions['B'].width = 6
    write_ws.column_dimensions['C'].width = 14
    write_ws.column_dimensions['D'].width = 8
    write_ws.column_dimensions['E'].width = 12

    write_ws.column_dimensions['F'].width = 14
    write_ws.column_dimensions['G'].width = 16
    write_ws.column_dimensions['H'].width = 10
    write_ws.column_dimensions['I'].width = 18

    write_ws.column_dimensions['J'].width = 12
    write_ws.column_dimensions['K'].width = 6
    write_ws.column_dimensions['L'].width = 12
    write_ws.column_dimensions['M'].width = 6
    write_ws.column_dimensions['N'].width = 12
    write_ws.column_dimensions['O'].width = 10


    # [Sub Title1]
    ws_sub_title1 = []
    ws_sub_title1.append({"locate":1, "text":"파일생성날짜"})
    ws_sub_title1.append({"locate":4, "text":"사용API"})
    ws_sub_title1.append({"locate":6, "text":"사용데이터셋"})
    ws_sub_title1.append({"locate":8, "text":"수행시간"})
    ws_sub_title1.append({"locate":10, "text":"성공개수"})
    ws_sub_title1.append({"locate":12, "text":"검사개수"})
    ws_sub_title1.append({"locate":14, "text":"성공률"})
    write_ws.merge_cells(f'A2:B2')

    # [Sub Title1 셀 서식]
    for i in range(0, 7):
        write_ws.cell(2, ws_sub_title1[i]['locate'], ws_sub_title1[i]["text"])
        write_ws.cell(2, ws_sub_title1[i]['locate']).alignment = Alignment(horizontal="center", vertical="center")
        write_ws.cell(2, ws_sub_title1[i]['locate']).font = Font(bold=True, color="FFFFFF")
        write_ws.cell(2, ws_sub_title1[i]['locate']).fill = PatternFill(fgColor="333333", fill_type="solid")

    # [Sub Title1 내용 입력]
    from datetime import datetime
    write_ws.cell(2, 3, datetime.today().strftime("%Y-%m-%d")).alignment = Alignment(horizontal="center")
    write_ws.cell(2, 5, "ETRI STT").alignment = Alignment(horizontal="center")
    write_ws.cell(2, 7, "한국인대화음성").alignment = Alignment(horizontal="center")
    write_ws.cell(2, 9, result_list[2]).alignment = Alignment(horizontal="right")
    write_ws.cell(2, 11, result_list[0]) .alignment = Alignment(horizontal="right")
    write_ws.cell(2, 13, result_list[1]) .alignment = Alignment(horizontal="right")
    write_ws.cell(2, 15, str(round(result_list[0]/result_list[1]*100)) + "%\n") .alignment = Alignment(horizontal="center")
    
    # [Sub Title2]
    ws_sub_title2 = []
    ws_sub_title2.append({"locate":1, "text":"No"})
    ws_sub_title2.append({"locate":2, "text":"음성데이터"})
    ws_sub_title2.append({"locate":4, "text":"정확도"})
    ws_sub_title2.append({"locate":6, "text":"원본 텍스트"})
    ws_sub_title2.append({"locate":10, "text":"전사 텍스트"})
    write_ws.merge_cells(f'B3:C3')
    write_ws.merge_cells(f'D3:E3')
    write_ws.merge_cells(f'F3:I3')
    write_ws.merge_cells(f'J3:O3')

    # [Sub Title2 셀 서식]
    for i in range(0, 5):
        write_ws.cell(3, ws_sub_title2[i]['locate'], ws_sub_title2[i]["text"])
        write_ws.cell(3, ws_sub_title2[i]['locate']).alignment = Alignment(horizontal="center", vertical="center")
        write_ws.cell(3, ws_sub_title2[i]['locate']).font = Font(bold=True, color="FFFFFF")
        write_ws.cell(3, ws_sub_title2[i]['locate']).fill = PatternFill(fgColor="333333", fill_type="solid")

    # [Sub Title2 내용 입력 : 정확도 95%미만 음성데이터 목록]
    count = 4
    for i in failed_list:
        write_ws.cell(count, 1, count-3)
        write_ws.merge_cells(f'B{count}:C{count}')
        write_ws.cell(count, 2, i['name'])
        write_ws.merge_cells(f'D{count}:E{count}')
        write_ws.cell(count, 4, round(float(i['ratio']), 2))
        write_ws.merge_cells(f'F{count}:I{count}')
        write_ws.cell(count, 6, i['origin_text'])
        write_ws.merge_cells(f'J{count}:O{count}')
        write_ws.cell(count, 10, i['trans_text'])
        
        write_ws.cell(count, 2).alignment = Alignment(horizontal="center")
        write_ws.cell(count, 4).alignment = Alignment(horizontal="right")
        write_ws.cell(count, 6).alignment = Alignment(horizontal="left")
        write_ws.cell(count, 10).alignment = Alignment(horizontal="left") 
        count += 1

    # [테두리 설정]
    THIN_BORDER = Border(Side('thin'),Side('thin'),Side('thin'),Side('thin'))
    for rng in write_ws[f'A1:O{count-1}']:
        for cell in rng:
            cell.border = THIN_BORDER

    # [파일 저장]
    try:
        write_wb.save('./data/test.xlsx')
        print("./data/test.xlsx 에 파일을 저장하였습니다.")
    except Exception as e:
        print("예외 :", e) 
        write_wb.save('./data/test_temp.xlsx')
        print("./data/test_temp.xlsx 에 파일을 저장하였습니다.")

def check_gantueo(text):
    # [간투어 표현 : 발성자가 다음 발성을 준비하기 위해서 소요되는 시간을 벌기 위해서 발성하는 것]
    # [간투어 체크 시작]
    import re
    text = re.sub("아니/", "아니", text)
    text = re.sub("저기/", "저기", text)
    text = re.sub("아/", "아", text)
    text = re.sub("그/", "그", text)
    text = re.sub("어/", "어", text)
    text = re.sub("음/", "음", text)
    text = re.sub("저/", "저", text)
    text = re.sub("으/", "으", text)
    text = re.sub("웅/", "웅", text)
    text = re.sub("에/", "에", text)
    # [간투어 체크 종료]
    return text

class SplitWavAudioMubin():
    def __init__(self, folder, filename):
        self.folder = folder
        self.filename = filename
        self.filepath = os.path.join(folder, filename + ".wav")
       
        self.audio = AudioSegment.from_wav(self.filepath)
        self.filelist = []
    
    def get_filelist(self):
        return self.filelist

    def get_duration(self):
        return self.audio.duration_seconds
    
    def single_split(self, from_min, to_min, split_filename):
        t1 = from_min * 20 * 1000
        t2 = to_min * 20 * 1000
        split_audio = self.audio[t1:t2]
        split_audio_export_path = os.path.join(self.folder, split_filename)
        split_audio.export(split_audio_export_path, format="wav")
        
    def multiple_split(self, min_per_split):
        total_mins = math.ceil(self.get_duration() / 20)
        for i in range(0, total_mins, min_per_split):
            split_fn = self.filename + '_' + str(i)
            self.filelist.append(split_fn)
            self.single_split(i, i+min_per_split, split_fn + ".wav")
            # print(str(i) + ' Done')
            # if i == total_mins - min_per_split:
            #     print('All splited successfully')

    def remove_splited_files(self):
        filelist = self.get_filelist()
        for i in range(0, len(filelist)):
            remove_path = os.path.join(self.folder, filelist[i] + ".wav")
            if(os.path.isfile(remove_path)):
                os.remove(remove_path)

class Worker (threading.Thread):
    def __init__(self, name, filename, folderpath, args1, args2):
        super().__init__()
        self.name = name
        self.filename = filename
        self.folderpath = folderpath
        self.args1 = args1
        self.args2 = args2

    def run(self):
      
        for i in range(self.args1, self.args2 + 1):
            # [For Adjustment file name]
            temp_name = ""
            ratio = 0.0
            origin_script = ""
            trans_script = ""

            if   i < 10:          temp_name = filename + '000' + str(i)
            elif 10 <= i < 100:   temp_name = filename + '00' + str(i)
            elif 100 <= i < 1000: temp_name = filename + '0' + str(i)

            # [Checking]
            # 20초 이상의 음성데이터는  ETRI_STT 한국어 인식 API를 사용할 수가 없다.
            # 정확히는 20초 이상의 음성데이터는 20초까지만 인식하고 나머지를 버린다.
            from scipy.io import wavfile

            audioFilePath = os.path.join(folderpath, temp_name + '.wav')
            fs, data = wavfile.read(audioFilePath)
            play_time = len(data)/fs

            if (play_time > 20):
                # 음성데이터가 잘림
                # 음성데이터 전사 전에 20초 단위로 잘라서 입력
                # 예:) 52초 음성데이터 -> 1~20, 21~40, 41~52  3개의 음성데이터로 분할한다.
                #    분할된 음성데이터 예시 hobby_00000149.wav
                #    hobby_00000149_0.wav
                #    hobby_00000149_1.wav
                #    hobby_00000149_2.wav
                # 분할된 음성데이터를 순서대로 분석하고 분석내용을 합친다.
                # 분할된 음성데이터들은 분석이 끝나면 삭제된다.
    
                # 분할 개체 생성
                split_wav = SplitWavAudioMubin(folderpath, temp_name)

                # 음성데이터 분할작업
                split_wav.multiple_split(min_per_split=1)

                # 분할된 각각의 음성데이터 전사처리 & 병합
                splited_filelist = split_wav.get_filelist()
                merge_trans_script = ""

                for i in range(0, len(splited_filelist)):
                    etri = ETRI_STT_API(folderpath, splited_filelist[i])
                    etri.one_click_splited()
                    merge_trans_script += etri.get_trans_script()

                # 분할된 데이터 제거
                split_wav.remove_splited_files()

                # ratio 측정
                m_etri = ETRI_STT_API(folderpath, temp_name)
                m_etri.set_trans_script(merge_trans_script)
                ratio = m_etri.analyze_script_difference(PRINT_RESULT_ON)
                origin_script = m_etri.get_origin_script()
                trans_script = m_etri.get_trans_script()

            else:
                etri = ETRI_STT_API(folderpath, temp_name)
                etri.one_click()
                ratio = etri.get_ratio()
                origin_script = etri.get_origin_script()
                trans_script = etri.get_trans_script()


            # ---- synchronization ----
            global count
            global failed_list
            lock.acquire()
            try:
                if ratio>=0.95:
                    count += 1
                else:
                    failed_list.append({'name':str(temp_name), 'ratio':str(ratio), 'origin_text':origin_script, 'trans_text':trans_script})

            finally:
                lock.release()


        


#######################################################################


# [main function]
# [This point is Program's starting point]
if __name__ == '__main__':
    # [For check Runtime]
    import time
    start = time.time()

    folderpath = os.path.join(os.path.dirname(__file__), 'aihub_data\\hobby_01\\001')
    filename = "hobby_0000"
    start_num = 2
    end_num = 5
    num_size = end_num - start_num + 1
    if (num_size<=0):       # [Error : Input range]
        exit()
    
    # ----- THREAD -----

    threads = []
    job_interval = num_size // NUM_OF_THREADS
    temp_point = 0
    for n in range(0, NUM_OF_THREADS):
        job_start_num = start_num + job_interval * n
        job_end_num = start_num + job_interval * (n + 1)
        temp_point = job_start_num

        thread = Worker(n, filename, folderpath, job_start_num, job_end_num)
        thread.start()
        threads.append(thread)
    job_start_num = start_num + job_interval * (NUM_OF_THREADS - 1)
    job_end_num = end_num + 1
    thread = Worker(NUM_OF_THREADS-1, filename, folderpath, job_start_num, job_end_num)
    thread.start()
    threads.append(thread)

    for thread in threads:
        thread.join()

        
                

    result_string = "--------------------------- | RESULT | ---------------------------\n"
    result_string += "success count : " + str(count) + " / " + str(num_size) + "\n"
    result_string += "success ratio : " + str(round(count/num_size*100)) + "%\n"
    result_string += "Runtime       : " + ShowRunTime(round(time.time() - start, 0))
    print(result_string)

    result_list = []
    result_list.append(count)
    result_list.append(num_size)
    result_list.append(ShowRunTime(round(time.time() - start, 0)))
    

    Output_result_as_EXCEL(failed_list, result_list)


