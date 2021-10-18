# 프로젝트 시작 2021년도 9월
# ------------------------------------------------------------------
# [ToDo 목록]
# 1. 함수 클래스화
# 
# ------------------------------------------------------------------
# [작성자]
#   조남규 (namkyu742@naver.com)
# ------------------------------------------------------------------
# [참고자료]
# 




#-*- coding:utf-8 -*-
from datetime import date
import urllib3
import json
import base64
import json
import os


class ETRI_STT_API():
    def __init__(self) -> None:
        self = ETRI_STT_API()

    def use_ETRI_API(file_name):
        openApiURL = "http://aiopen.etri.re.kr:8000/WiseASR/Recognition"
        accessKey = "198b2f86-c3a3-409c-b524-3f065eb25dd7"
        languageCode = "korean"

        audioFilePath = os.path.join(os.path.dirname(__file__), 'aihub_data', 'hobby_01', '001', file_name + '.wav')
        file = open(audioFilePath, "rb")
        audioContents = base64.b64encode(file.read()).decode("utf8")
        file.close()
        
        requestJson = {
            "access_key": accessKey,
            "argument": {
                "language_code": languageCode,
                "audio": audioContents
            }
        }
        
        http = urllib3.PoolManager()
        response = http.request(
            "POST",
            openApiURL,
            headers={"Content-Type": "application/json; charset=UTF-8"},
            body=json.dumps(requestJson)
        )

        data = json.loads(response.data.decode("utf-8", errors='ignore'))    
        return data['return_object']['recognized']


    def check_gantueo(text):
        # [간투어 표현 : 발성자가 다음 발성을 준비하기 위해서 소요되는 시간을 벌기 위해서 발성하는 것]
        # [간투어 체크 시작]
        import re
        text = re.sub("아니/", "아니", text)
        text = re.sub("저기/", "저기", text)
        text = re.sub("아/", "아", text)
        text = re.sub("그/", "그", text)
        text = re.sub("어/", "어", text)
        text = re.sub("음/", "음", text)
        text = re.sub("저/", "저", text)
        text = re.sub("으/", "으", text)
        text = re.sub("웅/", "웅", text)
        text = re.sub("에/", "에", text)
        # [간투어 체크 종료]
        return text

    def check_Double_Copying(text1, text2):
        # 이중전사 검사
        # 이중전사 : (철자전사)/(발음전사)
        import regex
        from difflib import SequenceMatcher

        list1 = regex.findall('\p{Hangul}+|\W|\d', text1)
        flag1 = False
        flag2 = False

        text_result = ""
        temp_text1 = ""
        temp_text2 = ""
            
        for i in range(0, len(list1)):
            if (list1[i] == '(' and flag1 == False and flag2 == False):
                flag1 = True
                continue
            if (list1[i] == ')' and flag1 == True and flag2 == False):
                flag2 = False
                continue
            if (list1[i] == '/' and flag1 == False and flag2 == False):
                continue
            if (list1[i] == '/' and flag1 == True and flag2 == False):
                flag1 = False
                flag2 = True
                continue
            if (list1[i] == '(' and flag1 == False and flag2 == True):
                flag1 = True
                continue
            if (list1[i] == ')' and flag1 == True and flag2 == True):
                flag1 = False
                flag2 = False
                # 비교문에 해당 문자열이 존재하는지 검사
                result1 = SequenceMatcher(None, temp_text1, text2).ratio()
                result2 = SequenceMatcher(None, temp_text2, text2).ratio()

                # 검사 결과에 따라 해당 문자열을 추가
                if (result1 > result2):
                    text_result += temp_text1
                elif (result1 < result2):
                    text_result += temp_text2
                else:
                    # 원문과 유사도 검사 결과가 둘이 동일한 경우, 문자열1을 채택
                    text_result += temp_text1 

                temp_text1 = ""
                temp_text2 = ""
                continue

            if (flag1 == True and flag2 == False):
                # 이중전사 1
                temp_text1 += list1[i]
            elif (flag1 == True and flag2 == True):
                # 이중전사 2
                temp_text2 += list1[i]

            # 그 외의 글자들은 그대로 복사
            if (flag1 == False):
                text_result += list1[i]        

        return text_result

    def check_Spacing_and_PunctuationMarks(text):
        # 띄어쓰기 제거
        # 문장부호 제거
        #   쉼표 ,
        #   느낌표 !
        #   물음표 ?
        #   마침표 .

        # 중복발성 기호 제거 '+'
        # 잡음 기호 제거 '*'

        text_result = ""
        for i in range(0, len(text)):
            if (text[i] == ' ' 
            or text[i] == '+' 
            or text[i] == '*'
            or text[i] == ',' 
            or text[i] == '!' 
            or text[i] == '?' 
            or text[i] == '.' ):
                continue        
            text_result = text_result + text[i]

        return text_result

    def text_difflib(file_name_origin, transScript):
        from difflib import SequenceMatcher
        import os
        
        file_name = os.path.join(
            os.path.dirname(__file__), 'aihub_data', 'hobby_01', '001', file_name_origin + '.txt')

        f1 = open(file_name, mode='rt', encoding='UTF8')

        text1 = f1.read()

        text2 = transScript
        
        # 영어 소문자를 대문자로 변경
        text2 = text2.upper()

        text3 = self.check_Double_Copying(text1, text2)
        
        text3 = check_Spacing_and_PunctuationMarks(text3)
        text4 = check_Spacing_and_PunctuationMarks(text2)
        ratio = SequenceMatcher(None, text3, text4).ratio()
        
        if (ratio < 0.95):
            print('[RESULT] :', file_name_origin)
            print('OriginalScript    :', text1)
            print('TranScript        :', text2)
            print('OriginalScript[T] :', text3)
            print('TranScript[T]     :', text4)
            print('RATIO :', ratio)
            print('----------------------------------------------------------')

        return ratio, text1

    def one_click(file_name):
        transScript = use_ETRI_API(file_name)
        ratio, origin_text = text_difflib(file_name, transScript)
        return ratio, origin_text, transScript

    def ShowRunTime(runningTime):
        run_second = int(runningTime)
        run_minute = 0
        run_hour = 0
        if (run_second > 3600):
            run_hour = int(run_second / 3600)
            run_second = run_second - (3600*run_hour)
        if (run_second > 60):
            run_minute = int(run_second / 60)
            run_second = run_second - (60*run_minute)

        run_time = ""
        if (run_hour > 0):
            run_time += str(run_hour) + "시간 "
        if (run_minute > 0):
            run_time += str(run_minute) + "분 "
        run_time += str(run_second) + "초"

        return run_time

    def Output_result_as_EXCEL(failed_list, result_list):
        from openpyxl import Workbook
        from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
        write_wb = Workbook()
        write_ws = write_wb.active


        write_ws.merge_cells('A1:M1')
        write_ws.cell(1, 1, "정확도 95% 이하 음성데이터 목록").alignment = Alignment(horizontal="center", vertical="center")
        write_ws.cell(1, 1).font = Font(size=15, bold=True)
        write_ws.row_dimensions[1].height = 40


        ws_sub_title1 = []
        ws_sub_title1.append({"locate":1, "text":"파일생성날짜"})
        ws_sub_title1.append({"locate":4, "text":"사용API"})
        ws_sub_title1.append({"locate":6, "text":"사용데이터셋"})
        ws_sub_title1.append({"locate":8, "text":"수행시간"})
        ws_sub_title1.append({"locate":10, "text":"성공개수"})
        ws_sub_title1.append({"locate":12, "text":"성공률"})
        write_ws.merge_cells(f'A2:B2')

        for i in range(0, 6):
            write_ws.cell(2, ws_sub_title1[i]['locate'], ws_sub_title1[i]["text"])
            write_ws.cell(2, ws_sub_title1[i]['locate']).alignment = Alignment(horizontal="center", vertical="center")
            write_ws.cell(2, ws_sub_title1[i]['locate']).font = Font(bold=True, color="FFFFFF")
            write_ws.cell(2, ws_sub_title1[i]['locate']).fill = PatternFill(fgColor="333333", fill_type="solid")

        from datetime import datetime
        write_ws.cell(2, 3, datetime.today().strftime("%Y-%m-%d")).alignment = Alignment(horizontal="center")
        write_ws.cell(2, 5, "ETRI STT").alignment = Alignment(horizontal="center")
        write_ws.cell(2, 7, "한국인대화음성").alignment = Alignment(horizontal="center")
        write_ws.cell(2, 9, result_list[2]).alignment = Alignment(horizontal="right")
        write_ws.cell(2, 11, result_list[0]) .alignment = Alignment(horizontal="right")
        write_ws.cell(2, 13, str(round(result_list[0]/result_list[1]*100)) + "%\n") .alignment = Alignment(horizontal="center")
        


        ws_sub_title2 = []
        ws_sub_title2.append({"locate":1, "text":"No"})
        ws_sub_title2.append({"locate":2, "text":"음성데이터"})
        ws_sub_title2.append({"locate":4, "text":"정확도"})
        ws_sub_title2.append({"locate":6, "text":"원본 텍스트"})
        ws_sub_title2.append({"locate":10, "text":"전사 텍스트"})
        
        write_ws.merge_cells(f'B3:C3')
        write_ws.merge_cells(f'D3:E3')
        write_ws.merge_cells(f'F3:I3')
        write_ws.merge_cells(f'J3:M3')

        for i in range(0, 5):
            write_ws.cell(3, ws_sub_title2[i]['locate'], ws_sub_title2[i]["text"])
            write_ws.cell(3, ws_sub_title2[i]['locate']).alignment = Alignment(horizontal="center", vertical="center")
            write_ws.cell(3, ws_sub_title2[i]['locate']).font = Font(bold=True, color="FFFFFF")
            write_ws.cell(3, ws_sub_title2[i]['locate']).fill = PatternFill(fgColor="333333", fill_type="solid")


        
        write_ws.column_dimensions['A'].width = 6
        write_ws.column_dimensions['B'].width = 6
        write_ws.column_dimensions['C'].width = 16
        write_ws.column_dimensions['D'].width = 8
        write_ws.column_dimensions['E'].width = 12

        write_ws.column_dimensions['F'].width = 12
        write_ws.column_dimensions['G'].width = 14
        write_ws.column_dimensions['H'].width = 8
        write_ws.column_dimensions['I'].width = 14

        write_ws.column_dimensions['J'].width = 12
        write_ws.column_dimensions['K'].width = 12
        write_ws.column_dimensions['L'].width = 12
        write_ws.column_dimensions['M'].width = 12


        count = 4
        for i in failed_list:
            write_ws.cell(count, 1, count-3)
            write_ws.merge_cells(f'B{count}:C{count}')
            write_ws.cell(count, 2, i['name'])
            write_ws.merge_cells(f'D{count}:E{count}')
            write_ws.cell(count, 4, round(float(i['ratio']), 2))
            write_ws.merge_cells(f'F{count}:I{count}')
            write_ws.cell(count, 6, i['origin_text'])
            write_ws.merge_cells(f'J{count}:M{count}')
            write_ws.cell(count, 10, i['trans_text'])
            
            write_ws.cell(count, 2).alignment = Alignment(horizontal="center")
            write_ws.cell(count, 4).alignment = Alignment(horizontal="right")
            write_ws.cell(count, 6).alignment = Alignment(horizontal="left")
            write_ws.cell(count, 10).alignment = Alignment(horizontal="left")
            count += 1

        THIN_BORDER = Border(Side('thin'),Side('thin'),Side('thin'),Side('thin'))
        for rng in write_ws[f'A1:M{count-1}']:
            for cell in rng:
                cell.border = THIN_BORDER

        write_wb.save('./data/test.xlsx')








# [main function]
# [This point is Program's starting point]
if __name__ == '__main__':
    # [For check Runtime]
    import time
    start = time.time()

    failed_list = []
    count = 0
    file_name = "hobby_0000"
    start_num = 201
    end_num = 220
    num_size = end_num - start_num + 1
    if (num_size<=0):
        # [Error : Input range]
        exit()

    for i in range(start_num, end_num + 1):
        # [For Adjustment file name]
        if i<10:
            temp_name = file_name + '000' + str(i)
        elif 10<=i<100:
            temp_name = file_name + '00' + str(i)
        elif 100<=i<1000:
            temp_name = file_name + '0' + str(i)

        # [Checking]
        # 20초 이상의 음성데이터는  ETRI_STT 한국어 인식 API를 사용할 수가 없다.
        # 정확히는 20초 이상의 음성데이터는 20초까지만 인식하고 나머지를 버린다.
        from scipy.io import wavfile

        audioFilePath = os.path.join(os.path.dirname(__file__), 'aihub_data', 'hobby_01', '001', temp_name + '.wav')
        fs, data = wavfile.read(audioFilePath)
        play_time = len(data)/fs
        print(play_time)

        if (play_time > 20):
            # 음성데이터가 잘림
            # 음성데이터 전사 전에 20초 단위로 잘라서 입력
            # 예:) 52초 음성데이터 -> 1~20, 21~40, 41~52  3가지 음성데이터로 분할한다.
            #    분할된 음성데이터 예시 hobby_00000149.wav
            #    hobby_00000149_1.wav
            #    hobby_00000149_2.wav
            #    hobby_00000149_3.wav
            # 분할된 음성데이터를 순서대로 분석하고 분석내용을 합친다.
            # 분할된 음성데이터들은 분석이 끝나면 삭제된다.
            print("음성데이터 분해")






        ratio, origin_text, trans_text = one_click(temp_name)
        if ratio>=0.95:
            count += 1
        else:
            failed_list.append({'name':str(temp_name), 'ratio':str(ratio), 'origin_text':origin_text, 'trans_text':trans_text})

    result_string = "----------------------- | RESULT | -----------------------\n"
    result_string += "success count : " + str(count) + " / " + str(num_size) + "\n"
    result_string += "success ratio : " + str(round(count/num_size*100)) + "%\n"
    result_string += "Runtime       :" + ShowRunTime(round(time.time() - start, 0))
    print(result_string)

    result_list = []
    result_list.append(count)
    result_list.append(num_size)
    result_list.append(ShowRunTime(round(time.time() - start, 0)))

    Output_result_as_EXCEL(failed_list, result_list)


