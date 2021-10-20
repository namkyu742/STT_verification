#-*- coding:utf-8 -*-

"""
프로젝트 시작 2021년도 9월
------------------------------------------------------------------
[작성자]
- 조남규 (namkyu742@naver.com)

------------------------------------------------------------------
[TODO 목록]
01. 완료 - 함수 클래스화
02. 완료 - 20초 이상의 음성데이터 분할처리
03. 완료 - 결과를 액셀파일로 출력
04. 완료 - ffmpeg 경고문제 해결하기
05. 액셀파일이 열려있을때 대처하기
06. 불가능 - 액셀파일에 원문과 전사문 다른점 강조표시
07. 불가능 - 쓰레드 사용해서 작업 분담
08. 완료 - 프로그램 로그 작성
09. 완료 - 변수, 함수 네이밍 법칙 정리
10. 액셀파일에 누락된 음성데이터 개수 넣기
11. API 선택기능

------------------------------------------------------------------
프로그램 동작 순서 구상안
- 폴더(또는 압축파일 - 추후 개발예정)를 지정하면 검사할 목록을 파악하고 메타데이터를 출력.
- 검사목록에서 데이터셋을 선택
- 검사 범위 설정 (전부 혹은 일부분)
- 검사
- 검사결과 출력
- 검사결과 파일로 출력
- 검사목록으로 돌아감
- 프로그램 종료 선택 시 프로그램 종료

- * 프로그램 실행 중에는 수행한 작업에 대한 로그를 작성 
-   (https://wikidocs.net/123324)
-   (https://minimin2.tistory.com/41)

------------------------------------------------------------------
프로그램의 목적
- 음성데이터의 의미적 정확성 검사
- 의미정확성이란? 
    - 데이터에 부착한 라벨링 값이 실제 참값과 일치하도록 데이터의 의미적인 정확성을 확보하는 것.
    - 학습모델의 성능을 좌우하는 품질 요소로 매우 중요하며, 이 때문에 2인 이상의 교차검증, 전수검사를 원칙으로 한다.
    - 검사 대상 라벨링 내용
        - 음성데이터 : 전사 텍스트
        - 텍스트데이터 : 내용요약, 번역, 질의응답, 말뭉치 태깅
    - 정확도 : 어노테이션과 참값 간의 중첩율 (예: IoU, ROUGE, BLEU, F1, EM 등)

------------------------------------------------------------------
음성인식 전사텍스트 인식률 개선 방안
- 한글 맞춤법 라이브러리 사용
  -  https://blog.naver.com/PostView.naver?blogId=dsz08082&logNo=222317249656&redirect=Dlog&widgetTypeCall=true&directAccess=false

  -  clone https://github.com/ssut/py-hanspell
  -  cd py-hanspell
  -  python setup.py install
  -  네이버 맞춤법 검사기 API를 사용하는 라이브러리인데 별로다.
  -  사용 안함

------------------------------------------------------------------
GUI 설계 방안
- GUI프로그램? WEB page? CUI로? UI없이 터미널에서 직접 실행?
    - C# 사용을 고려중
- 서버가 필요한가
    - 데이터의 저장?

------------------------------------------------------------------
[참고자료]
날짜시간 가져오기 (https://yujuwon.tistory.com/entry/%ED%98%84%EC%9E%AC-%EB%82%A0%EC%A7%9C-%EA%B0%80%EC%A0%B8%EC%98%A4%EA%B8%B0)
예외처리 (https://wikidocs.net/30#try-finally)
음성데이터 분할 (https://stackoverflow.com/questions/37999150/how-to-split-a-wav-file-into-multiple-wav-files)
os 라이브러리 사용법 (https://webisfree.com/2018-03-16/python-%ED%8C%8C%EC%9D%BC-%EB%B0%8F-%EB%94%94%EB%A0%89%ED%86%A0%EB%A6%AC-%EC%82%AD%EC%A0%9C%ED%95%98%EB%8A%94-%EB%B0%A9%EB%B2%95)
os 라이브러리 사용법2 (https://yganalyst.github.io/data_handling/memo_1/)
실행파일 만들기1 (https://wikidocs.net/21952)
실행파일 만들기2 (https://m.blog.naver.com/smilewhj/221070338758)
파일 존재여부검사 (https://wikidocs.net/14304)
"""

# ----------------------- [imports] -----------------------
import urllib3
import json
import base64
import os
from pydub import AudioSegment
import math
import time
import requests

# ------------------ [CONSTANT VARIALBE] ------------------
# API와 데이터셋 설정, 기준정확도에 대한 부분은 추후 개발 예정
USED_API = "ETRI STT"
USED_DATA_SET = "한국인대화음성"
TARGET_RATIO = 0.95
# FLAGS_FOR_EXCEPTION
TEST_FOLDERPATH = 0b000001
TEST_FILENAME   = 0b000010
TEST_RANGE      = 0b000100

PROGRAM_PATH = os.path.dirname(__file__)

# ------------------- [for CODE LOGGING] -------------------
# 참고 링크 (https://wikidocs.net/123324)
from logging.config import dictConfig
import logging
dictConfig({
    'version': 1,
    'formatters': {
        'default': {
            'format': '[%(asctime)s] %(message)s',
        }
    },
    'handlers': {
        'file': {
            'level': 'DEBUG',
            'class': 'logging.FileHandler',
            'filename': PROGRAM_PATH + '/data/debug.log',
            'formatter': 'default',
        },
    },
    'root': {
        'level': 'DEBUG',
        'handlers': ['file']
    }
})

# ------------------ [Speech To Text API] ------------------
class useSTT_API():
    """
    음성인식 엔진을 사용하여 음성인식정확도를 검증하기 위한 클래스

    사용가능한 음성인식엔진:
        - ETRI STT Engine
        - Google STT Engine
        - Kakao STT Engine

    Args:
        - folderPath : 파일의 경로
        - fileName : 파일의 이름
        - apiAccessKey : API를 사용하기 위해 인증받은 Key
    
    경로와 이름을 통해 wav파일과 txt파일을 읽고,
    wav파일을 API를 통해 전사한 후 txt파일과 비교하여 정확도를 측정한 후 반환한다.
    """
    
    def __init__(self, folderPath:str, fileName:str, apiAccessKey:dict, selectedApi:str):
        self.folderPath = folderPath        # 음성데이터의 폴더 경로
        self.fileName = fileName            # 음성데이터의 파일 이름(넘버링 앞부분)
        self.transScript = ""               # 전사 텍스트
        self.originScript = ""              # 라벨링 텍스트
        self.ratio = 0.0                    # 정확도
        self.selectedApi = selectedApi
        self.selectedApiAccessKey = "None"
        if (selectedApi == "ETRI") : self.selectedApiAccessKey = apiAccessKey['ETRI']
        elif (selectedApi == "Kakao"): self.selectedApiAccessKey = apiAccessKey['Kakao']
        elif (selectedApi == "Google"): self.selectedApiAccessKey = apiAccessKey['Google']
        else:
            logging.info("Incorrect API selection : " + selectedApi)
            print("Incorrect API selection : " + selectedApi)
            exit()

        filePath = os.path.join(self.folderPath, self.fileName + ".txt")
        try:
            scriptFile = open(filePath, mode='rt', encoding='UTF8')
            logging.info("Success to open script file : " + filePath)
            self.originScript = scriptFile.read()
        except Exception as e:
            logging.info("Failed to open script file : " + filePath)
            print("exception: ", e)
        finally:
            scriptFile.close()

    def getApiAccessKey(self):
        return self.selectedApiAccessKey
        
    def getTransScript(self)->str:
        """전사 텍스트를 반환한다."""
        return self.transScript

    def setTransScript(self, transScript:str):
        """전사 텍스트를 인스턴트변수에 저장한다."""
        self.transScript = transScript

    def getRatio(self)->float:
        """정확도를 반환한다."""
        return self.ratio
        
    def getFilePath(self)->str:
        """음성데이터의 파일이름과 확장자를 포함한 파일경로를 반환한다."""
        return os.path.join(self.folderPath, self.fileName + ".wav")

    def getOriginScript(self)->str:
        """라벨링 텍스트를 반환한다."""
        return self.originScript

    def etriSTTEngine(self)->str:
        """API를 호출한다."""
        openApiURL = "http://aiopen.etri.re.kr:8000/WiseASR/Recognition"
        languageCode = "korean"

        audiofilePath = self.getFilePath()
        try:
            file = open(audiofilePath, "rb")
        except Exception as e:
            print("Exception:", e)
            logging.info("occur Exception : " + e)
            exit()
        logging.info("open audio file : " + audiofilePath)
        audioContents = base64.b64encode(file.read()).decode("utf8")
        file.close()
        
        requestJson = {
            "access_key": self.getApiAccessKey(),
            "argument": {
                "language_code": languageCode,
                "audio": audioContents
            }
        }
        
        http = urllib3.PoolManager()
        response = http.request(
            "POST",
            openApiURL,
            headers={"Content-Type": "application/json; charset=UTF-8"},
            body=json.dumps(requestJson)
        )
        logging.info("request etri stt")

        data = json.loads(response.data.decode("utf-8", errors='ignore'))
        return data['return_object']['recognized']

    def kakaoSTTEngine(self)->str:
        kakao_speech_url = "https://kakaoi-newtone-openapi.kakao.com/v1/recognize"

        headers = {
            "Content-Type": "application/octet-stream",
            "X-DSS-Service": "DICTATION",
            "Authorization": "KakaoAK " + self.getApiAccessKey(),
        }

        audiofilePath = self.getFilePath()
        with open(audiofilePath, 'rb') as fp:
            audio = fp.read()

        res = requests.post(kakao_speech_url, headers=headers, data=audio)
        result_json_string = res.text[res.text.index('{"type":"finalResult"'):res.text.rindex('}')+1]
        result = json.loads(result_json_string)
        
        return result['value']
        
    def googleSTTEngine(self)->str:
        import io
        from google.cloud import speech

        client = speech.SpeechClient()
        audiofilePath = self.getFilePath()

        # Loads the audio into memory
        with io.open(audiofilePath, 'rb') as audio_file:
            content = audio_file.read()
            audio = speech.RecognitionAudio(content=content)

        config = speech.RecognitionConfig(
            encoding=speech.RecognitionConfig.AudioEncoding.LINEAR16,
            sample_rate_hertz=16000, #44000
            # audio_channel_count=2,
            # speech_contexts=[speech.types.SpeechContext(
            # phrases=['hi', 'good afternoon']
            # )],
            language_code='ko-KR')

        # Detects speech in the audio file
        response = client.recognize(config=config, audio=audio)

        for result in response.results:
            result_return = format(result.alternatives[0].transcript)
                        
        return result_return

    def chooseSTTEngine(self):
        if (self.selectedApi == "ETRI"): self.transScript = self.etriSTTEngine()
        elif (self.selectedApi == "Google") : self.transScript = self.googleSTTEngine()
        elif (self.selectedApi == "Kakao") : self.transScript = self.kakaoSTTEngine()
        else:
            logging.info("Incorrect API selection")
            exit()
      
    def checkDoubleCopying(self):
        """
        이중전사 (철자전사)/(발음전사) 검사
        라벨링텍스트의 철자전사와 발음전사 중 어느쪽이 전사텍스트에 가까운지 검사하여 선택한다.
        동시에 이중전사 기호('(', ')', '/')를 문자열에서 제거한다.
        또한, 이 단계에서 간투어표현을 위해 사용되는 기호 '/'도 같이 제거된다.
        """
        import regex
        from difflib import SequenceMatcher
        
        transScript = self.transScript
        checkingString = regex.findall('\p{Hangul}+|\W|\d', self.getOriginScript()) 
        """
        regex.findall()은 정규식과 매치되는 모든 문자열을 리스트형식으로 리턴한다.
        위 정규식의 의미 : 
            - \p{Hangul}+   : 한글
            - \W            : 단어 문자가 아닌 모든 문자
            - \d            : 모든 유니코드 십진수
        """
        # flag 변수 초기화
        flag1 = False       # 괄호가 열려있으면 True
        flag2 = False       # 두번째 괄호(발음전사)일때 True

        resultScript = ""   # 검사 결과를 기록할 문자열
        tempScript1 = ""    # 철자전사를 기록할 문자열
        tempScript2 = ""    # 발음전사를 기록할 문자열
            
        for i in range(0, len(checkingString)):
            if (checkingString[i] == '(' and flag1 == False and flag2 == False):
                flag1 = True
                continue
            if (checkingString[i] == ')' and flag1 == True and flag2 == False):
                flag2 = False
                continue
            if (checkingString[i] == '/' and flag1 == False and flag2 == False):
                continue
            if (checkingString[i] == '/' and flag1 == True and flag2 == False):
                flag1 = False
                flag2 = True
                continue
            if (checkingString[i] == '(' and flag1 == False and flag2 == True):
                flag1 = True
                continue
            if (checkingString[i] == ')' and flag1 == True and flag2 == True):
                flag1 = False
                flag2 = False
                # 전사문에 해당 문자열이 존재하는지 검사
                result1 = SequenceMatcher(None, tempScript1, transScript).ratio()
                result2 = SequenceMatcher(None, tempScript2, transScript).ratio()
                
                # 검사 결과에 따라 해당 문자열을 추가
                # 원문과 유사도 검사 결과가 둘이 동일한 경우, 철자전사를 채택
                if (result1 > result2): resultScript += tempScript1
                elif (result1 < result2): resultScript += tempScript2
                else: resultScript += tempScript1 

                # tempScript 초기화
                tempScript1 = ""
                tempScript2 = ""
                continue

            if (flag1 == True and flag2 == False):  # tempScript1에 철자전사 기록
                tempScript1 += checkingString[i]
            elif (flag1 == True and flag2 == True): # tempScript2에 발음전사 기록
                tempScript2 += checkingString[i]
            else:       # 이중전사에 해당되지 않는 글자는 그대로 resultScript에 기록
                resultScript += checkingString[i]        

        return resultScript

    def checkSpacingandPunctuationMarks(self, script):
        """
        다음의 기호들을 제거
            - 공백
            - 쉼표          ,
            - 느낌표        !
            - 물음표        ?
            - 마침표        .
            - 중복발성 기호 +
            - 잡음 기호     *
        """
        resultScript = ""
        for i in range(0, len(script)):
            if (script[i] == ' ' 
            or script[i] == '+' 
            or script[i] == '*'
            or script[i] == ',' 
            or script[i] == '!' 
            or script[i] == '?' 
            or script[i] == '.' ):
                continue        
            resultScript = resultScript + script[i]
        return resultScript

    def analyzeScriptDifference(self, option):
        """
        라벨링텍스트와 전사텍스트를 비교하여 음성인식의 의미적정확도를 측정한다.
            - option : 콘솔에 비교 결과를 출력할 것인지 선택 (단, 기준정확도 미달 데이터만 출력됨)
        사용하는 함수는 jellyfish 라이브러리의 jaro_winkler_similarity함수.
            - jaro거리 : [0,1]에서 부동 소수점 응답을 제공하는 문자열 편집 거리
            - 0은 완전히 다른 두 문자열을, 1은 동일한 문자열을 나타냄
            - jaro-winkler 위키 : https://en.wikipedia.org/wiki/Jaro%E2%80%93Winkler_distance
        """
        import jellyfish
    
        text1 = self.getOriginScript()
        text2 = self.transScript
        
        text2 = text2.upper()   # 영어 소문자를 대문자로 변경
        text3 = self.checkDoubleCopying()
        
        text3 = self.checkSpacingandPunctuationMarks(text3)
        text4 = self.checkSpacingandPunctuationMarks(text2)
        
        ratio = jellyfish.jaro_winkler_similarity(text3, text4)
        if (option == 1):
            if (ratio < TARGET_RATIO):
                print('[RESULT] :', self.fileName)
                print('OriginalScript    :', text1)
                print('TranScript        :', text2)
                print('OriginalScript[T] :', text3)
                print('TranScript[T]     :', text4)
                print('RATIO :', ratio)
                print('------------------------------------------------------------------')

        logging.info(self.fileName + "'s ratio : " + str(ratio))
        return ratio

    def oneClick(self):
        """
        음성데이터가 20초 미만일 경우. 인스턴스 변수에 의미적정확도를 저장
        """
        self.chooseSTTEngine()
        self.ratio = self.analyzeScriptDifference(1) # 매개변수 1이면 결과를 콘솔에 출력

    def oneClickSplited(self):
        """
        음성데이터가 20초 이상일 경우. 분할된 각각의 음성데이터에 대해 API를 사용하여 전사텍스트를 추출
        """
        self.chooseSTTEngine()
        
def ShowRunTime(runningTime):
    """
    초 단위로 주어지는 수행시간을 "시간", "분", "초" 포멧에 맞추어 반환
    """
    runSecond = int(runningTime)
    runMinute = 0
    runHour = 0
    if (runSecond > 3600):
        runHour = int(runSecond / 3600)
        runSecond = runSecond - (3600 * runHour)
    if (runSecond > 60):
        runMinute = int(runSecond / 60)
        runSecond = runSecond - (60 * runMinute)

    runTime = ""
    if (runHour > 0):
        runTime += str(runHour) + "시간 "
    if (runMinute > 0):
        runTime += str(runMinute) + "분 "
    runTime += str(runSecond) + "초"

    return runTime

def OutputResultAsEXCEL(failedList, resultList, missingList):
    """
    xlsx파일로 의미적정확도가 기준을 미달한 음성데이터에 대해 기록.
    기록되는 항목
        - 파일생성날짜, 사용된 API, 사용된 데이터셋, 수행시간, 성공개수, 검사개수, 성공률
        - 파일 넘버, 음성데이터 파일명, 정확도, 라벨링텍스트, 전사텍스트
    """
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
    writeWB = Workbook()
    writeWS = writeWB.active

    # [Title]
    writeWS.merge_cells('A1:O1')
    writeWS.cell(1, 1, "정확도 95% 이하 음성데이터 목록").alignment = Alignment(horizontal="center", vertical="center")
    writeWS.cell(1, 1).font = Font(size=15, bold=True)
    writeWS.row_dimensions[1].height = 40

    # [액셀 열 간격 조정]
    writeWS.column_dimensions['A'].width = 6
    writeWS.column_dimensions['B'].width = 6
    writeWS.column_dimensions['C'].width = 14
    writeWS.column_dimensions['D'].width = 8
    writeWS.column_dimensions['E'].width = 12

    writeWS.column_dimensions['F'].width = 14
    writeWS.column_dimensions['G'].width = 16
    writeWS.column_dimensions['H'].width = 10
    writeWS.column_dimensions['I'].width = 18

    writeWS.column_dimensions['J'].width = 12
    writeWS.column_dimensions['K'].width = 6
    writeWS.column_dimensions['L'].width = 12
    writeWS.column_dimensions['M'].width = 6
    writeWS.column_dimensions['N'].width = 12
    writeWS.column_dimensions['O'].width = 10

    # [Sub Title1]
    wsSubTitle1 = []
    wsSubTitle1.append({"locate":1, "text":"파일생성날짜"})
    wsSubTitle1.append({"locate":4, "text":"사용API"})
    wsSubTitle1.append({"locate":6, "text":"사용데이터셋"})
    wsSubTitle1.append({"locate":8, "text":"수행시간"})
    wsSubTitle1.append({"locate":10, "text":"성공개수"})
    wsSubTitle1.append({"locate":12, "text":"검사개수"})
    wsSubTitle1.append({"locate":14, "text":"성공률"})
    writeWS.merge_cells(f'A2:B2')

    # [Sub Title1 셀 서식]
    for i in range(0, 7):
        writeWS.cell(2, wsSubTitle1[i]['locate'], wsSubTitle1[i]["text"])
        writeWS.cell(2, wsSubTitle1[i]['locate']).alignment = Alignment(horizontal="center", vertical="center")
        writeWS.cell(2, wsSubTitle1[i]['locate']).font = Font(bold=True, color="FFFFFF")
        writeWS.cell(2, wsSubTitle1[i]['locate']).fill = PatternFill(fgColor="333333", fill_type="solid")

    # [Sub Title1 내용 입력]
    from datetime import datetime
    writeWS.cell(2, 3, datetime.today().strftime("%Y-%m-%d")).alignment = Alignment(horizontal="center")
    writeWS.cell(2, 5, USED_API).alignment = Alignment(horizontal="center")
    writeWS.cell(2, 7, USED_DATA_SET).alignment = Alignment(horizontal="center")
    writeWS.cell(2, 9, resultList[2]).alignment = Alignment(horizontal="right")
    writeWS.cell(2, 11, resultList[0]) .alignment = Alignment(horizontal="right")
    writeWS.cell(2, 13, resultList[1]) .alignment = Alignment(horizontal="right")
    writeWS.cell(2, 15, str(round(resultList[0]/resultList[1]*100)) + "%\n") .alignment = Alignment(horizontal="center")
    
    # [Sub Title2]
    wsSubTitle2 = []
    wsSubTitle2.append({"locate":1, "text":"No"})
    wsSubTitle2.append({"locate":2, "text":"음성데이터"})
    wsSubTitle2.append({"locate":4, "text":"정확도"})
    wsSubTitle2.append({"locate":6, "text":"라벨링 텍스트"})
    wsSubTitle2.append({"locate":10, "text":"전사 텍스트"})
    writeWS.merge_cells(f'B3:C3')
    writeWS.merge_cells(f'D3:E3')
    writeWS.merge_cells(f'F3:I3')
    writeWS.merge_cells(f'J3:O3')

    # [Sub Title2 셀 서식]
    for i in range(0, 5):
        writeWS.cell(3, wsSubTitle2[i]['locate'], wsSubTitle2[i]["text"])
        writeWS.cell(3, wsSubTitle2[i]['locate']).alignment = Alignment(horizontal="center", vertical="center")
        writeWS.cell(3, wsSubTitle2[i]['locate']).font = Font(bold=True, color="FFFFFF")
        writeWS.cell(3, wsSubTitle2[i]['locate']).fill = PatternFill(fgColor="333333", fill_type="solid")

    # [Sub Title2 내용 입력 : 정확도 95%미만 음성데이터 목록]
    count = 4
    for i in failedList:
        writeWS.cell(count, 1, count-3)
        writeWS.merge_cells(f'B{count}:C{count}')
        writeWS.cell(count, 2, i['name'])
        writeWS.merge_cells(f'D{count}:E{count}')
        writeWS.cell(count, 4, round(float(i['ratio']), 2))
        writeWS.merge_cells(f'F{count}:I{count}')
        writeWS.cell(count, 6, i['origin_text'])
        writeWS.merge_cells(f'J{count}:O{count}')
        writeWS.cell(count, 10, i['trans_text'])
        
        writeWS.cell(count, 2).alignment = Alignment(horizontal="center")
        writeWS.cell(count, 4).alignment = Alignment(horizontal="right")
        writeWS.cell(count, 6).alignment = Alignment(horizontal="left")
        writeWS.cell(count, 10).alignment = Alignment(horizontal="left") 
        count += 1

    # [Sub Title2 내용 입력 : 파일이 누락된 음성데이터 목록]
    for i in missingList:
        writeWS.cell(count, 1, count-3)
        writeWS.merge_cells(f'B{count}:C{count}')
        writeWS.cell(count, 2, i['name'])
        writeWS.merge_cells(f'D{count}:E{count}')
        writeWS.cell(count, 4, i['ratio'])
        writeWS.merge_cells(f'F{count}:I{count}')
        writeWS.cell(count, 6, i['origin_text'])
        writeWS.merge_cells(f'J{count}:O{count}')
        writeWS.cell(count, 10, i['trans_text'])
        
        for j in range(1, 16):
            writeWS.cell(count, j).fill = PatternFill(fgColor="BFBFBF", fill_type="solid")
        
        writeWS.cell(count, 2).alignment = Alignment(horizontal="center")
        writeWS.cell(count, 4).alignment = Alignment(horizontal="center")
        writeWS.cell(count, 6).alignment = Alignment(horizontal="center")
        writeWS.cell(count, 10).alignment = Alignment(horizontal="center") 
        count += 1

    # [테두리 설정]
    THIN_BORDER = Border(Side('thin'),Side('thin'),Side('thin'),Side('thin'))
    for rng in writeWS[f'A1:O{count-1}']:
        for cell in rng:
            cell.border = THIN_BORDER

    # [파일 저장]
    from datetime import datetime
    time = datetime.today().strftime("%Y-%m-%d_%H-%M-%S")
    savePath = os.path.join(PROGRAM_PATH, 'data', 'failed_list', 'result_' + time + '.xlsx')

    try:
        writeWB.save(savePath)
    except Exception as e:
        """
        TODO 예외처리 실패
        예외처리를 하였음에도 특정상황에서 에러가 발생함
            - savePath가 ""(공백)일 경우
        지정한 경로 이외의 임의의 문자가 savePath에 저장될 경우
            - ex) "abc" -> 파이썬 코드가 존재하는 폴더에 abc 파일이 생성됨 -> 액셀파일로 열면 내용 확인 가능(확장자 없음)
        """
        print(savePath, " 에 파일을 저장하지 못하였습니다.")
        print("[Error] ", e)
        logging.info("Failed to save excel file to [" + savePath + "]")
        logging.info("[Error] ", e)
    else:
        print(savePath, " 에 파일을 저장하였습니다.")
        logging.info("Success to save excel file to [" + savePath + "]")


class SplitWavAudioMubin():
    """
    음성데이터를 분할하기 위한 클래스
    출처 : https://stackoverflow.com/questions/37999150/how-to-split-a-wav-file-into-multiple-wav-files
    
    """
    def __init__(self, folder, fileName):
        self.folder = folder
        self.fileName = fileName
        self.filePath = os.path.join(folder, fileName + ".wav")
       
        self.audio = AudioSegment.from_wav(self.filePath)
        self.filelist = []
    
    def getFilelist(self):
        return self.filelist

    def getDuration(self):
        return self.audio.duration_seconds
    
    def singleSplit(self, from_min, to_min, split_fileName):
        t1 = from_min * 20 * 1000
        t2 = to_min * 20 * 1000
        split_audio = self.audio[t1:t2]
        split_audio_export_path = os.path.join(self.folder, split_fileName)
        split_audio.export(split_audio_export_path, format="wav")
        
    def multiple_split(self, min_per_split):
        total_mins = math.ceil(self.getDuration() / 20)
        for i in range(0, total_mins, min_per_split):
            split_fn = self.fileName + '_' + str(i)
            self.filelist.append(split_fn)
            self.singleSplit(i, i+min_per_split, split_fn + ".wav")
            # print(str(i) + ' Done')
            # if i == total_mins - min_per_split:
            #     print('All splited successfully')

    def remove_splited_files(self):
        filelist = self.getFilelist()
        for i in range(0, len(filelist)):
            remove_path = os.path.join(self.folder, filelist[i] + ".wav")
            if(os.path.isfile(remove_path)):
                os.remove(remove_path)

def examinationSTT(p_start:int, p_end:int, folderPath:str, fileName:str, selectedApi:str):
    """
    Args:
        - p_start: 검사 시작점
        - p_end: 검사 종료점
        - folderPath: 파일이 존재하는 폴더 경로
        - fileName: 파일 이름
        - selectedApi: 선택한 API

    파일이름 + 시작점(빈 자리수는 0으로 채움)에서 파일이름 + 끝점(빈 자리수는 0으로 채움)까지 검토

    음성데이터가 20초 이상일 경우 20초 단위로 분할하여 개별 텍스트 전사 후 텍스트를 합병하여 의미적정확도 검사
    """
    # [For check Runtime]
    start = time.time()

    failedList = []        # 정확도 기준 미달 음성데이터 목록
    missingList = []       # 파일이 존재하지 않는 음성데이터 목록
    count = 0              # 정확도 기준 달성 음성데이터 개수
    noFileCount = 0        # 파일이 존재하지 않는 음성데이터 개수

    API_ACCESS_KEY = {"Name":"Key"}
    API_ACCESS_KEY["ETRI"] = "198b2f86-c3a3-409c-b524-3f065eb25dd7" # ETRI API 접근 Key
    API_ACCESS_KEY["Kakao"] = "104c68aae23c3ffae3f622d01c9165e1"    # Kakao API 접근 Key
    API_ACCESS_KEY["Google"] = "None"                               # Google API Key 없음

    startNum = p_start     # 검사 시작 번호
    endNum = p_end         # 검사 끝 번호
    numSize = endNum - startNum + 1  # 작업량    
    if (numSize <= 0):       # [Error : Input range]
        logging.info("It's out of the range of files.")
        print("It's out of the range of files.")
        exit()

    for i in range(startNum, endNum + 1):
        # [Numbering file name]
        if i<10:            tempName = fileName + '000' + str(i)
        elif 10<=i<100:     tempName = fileName + '00' + str(i)
        elif 100<=i<1000:   tempName = fileName + '0' + str(i)

        # [Checking]
        # 20초 이상의 음성데이터는 ETRI 한국어 음성인식 API를 사용할 수가 없다.
        # ETRI 한국어 음성인식 API는 20초 이상의 음성데이터의 경우 20초까지만 인식하고 나머지를 버린다.
        from scipy.io import wavfile

        audiofilePath = os.path.join(folderPath, tempName + '.wav')
        # 파일이 존재하지 않을 경우 count 증가시키고 다음 파일로 넘어감
        if (os.path.isfile(audiofilePath) == False):
            logging.info("Path ["+audiofilePath+"] is no file")
            print("There is no file")
            noFileCount += 1
            missingList.append({
                'name':str(tempName), 
                'ratio':"no data", 
                'origin_text':"no data", 
                'trans_text':"no data"
                })
            continue

        fs, data = wavfile.read(audiofilePath)
        playTime = len(data)/fs

        if (playTime > 20):
            # 음성데이터의 길이가 20초를 초과할 경우 음성데이터 전사 전에 20초 단위로 잘라서 입력
            # 예:) 52초 음성데이터 -> 1~20, 21~40, 41~52  3개의 음성데이터로 분할한다.
            #    분할된 음성데이터 예시 hobby_00000149.wav
            #    hobby_00000149_0.wav
            #    hobby_00000149_1.wav
            #    hobby_00000149_2.wav
            # 분할된 음성데이터를 순서대로 전사하고 전사텍스트를 합친다.
            # 분할된 음성데이터들은 분석이 끝나면 삭제된다.
  
            # 분할 개체 생성
            splitWAV = SplitWavAudioMubin(folderPath, tempName)

            # 음성데이터 분할작업
            splitWAV.multiple_split(min_per_split=1)

            # 분할된 각각의 음성데이터 전사처리 & 병합
            splitedFileList = splitWAV.getFilelist()
            mergeTransScript = ""      # 분할된 전사텍스트 병합

            for j in range(0, len(splitedFileList)):
                sttAPI = useSTT_API(folderPath, splitedFileList[j], API_ACCESS_KEY, selectedApi)
                sttAPI.oneClickSplited()
                mergeTransScript += sttAPI.getTransScript()

            # 분할된 데이터 제거
            splitWAV.remove_splited_files()

            # ratio 측정
            m_sttAPI = useSTT_API(folderPath, tempName, API_ACCESS_KEY, selectedApi)
            m_sttAPI.setTransScript(mergeTransScript)
            ratio = m_sttAPI.analyzeScriptDifference(1)

            # 데이터 평가
            if (ratio >= TARGET_RATIO):
                count += 1
            else:
                failedList.append({
                    'name':str(tempName), 
                    'ratio':str(ratio), 
                    'origin_text':m_sttAPI.getOriginScript(), 
                    'trans_text':m_sttAPI.getTransScript()
                    })

        else:
            sttAPI = useSTT_API(folderPath, tempName, API_ACCESS_KEY, selectedApi)
            sttAPI.oneClick()
            ratio = sttAPI.getRatio()

            # 데이터 평가
            if (ratio >= TARGET_RATIO):
                count += 1
            else:
                failedList.append({
                    'name':str(tempName), 
                    'ratio':str(ratio), 
                    'origin_text':sttAPI.getOriginScript(), 
                    'trans_text':sttAPI.getTransScript()
                    })

    # 콘솔에 데이터 출력
    # 성공개수, 성공률, 없는파일개수, 수행시간
    resultString = "--------------------------- | RESULT | ---------------------------\n"
    resultString += "success count : " + str(count) + " / " + str(numSize-noFileCount) + "\n"
    resultString += "missing count : " + str(noFileCount) + "\n"
    resultString += "success ratio : " + str(round(count/(numSize-noFileCount)*100)) + "%\n"
    resultString += "Runtime       : " + ShowRunTime(round(time.time() - start, 0))
    print(resultString)

    # 액셀파일용 데이터
    resultList = []
    resultList.append(count)
    resultList.append(numSize-noFileCount)
    resultList.append(ShowRunTime(round(time.time() - start, 0)))
    resultList.append(noFileCount)
    OutputResultAsEXCEL(failedList, resultList, missingList)

def showMenu():
    """
    프로그램 메뉴를 출력하는 함수
    """
    print("--- MENU ---")
    print("1. RUN")
    print("2. ")
    print("3. ")
    print("0. EXIT")


def Run():
    """
    프로그램 실행
    - 시작번호, 끝번호, 파일경로, 파일이름, 사용할 API가 설정되어있음.
    """
    start = 7
    end = 7
    folderPath = os.path.join(PROGRAM_PATH, 'aihub_data\\hobby_01\\001')
    fileName = "hobby_0000"
    selectedApi = "ETRI"
    # selectedApi = "Google"
    # selectedApi = "Kakao"

    logging.info("-------------------------- Program start --------------------------")
    logging.info("path : " + folderPath)
    logging.info("fileName : " + fileName + " / range : " + str(start) + " ~ " + str(end))
    examinationSTT(start, end, folderPath, fileName, selectedApi)
    examinationSTT(start, end, folderPath, fileName, "Google")
    examinationSTT(start, end, folderPath, fileName, "Kakao")
    logging.info("--------------------------- Program end ---------------------------")


def exceptionTest(option):
    """
    예외 상황, 에러발생 등을 테스트
    """
    start = 7
    end = 7
    folderPath = os.path.join(PROGRAM_PATH, 'aihub_data\\hobby_01\\001')
    fileName = "hobby_0000"

    if (option & TEST_FOLDERPATH == TEST_FOLDERPATH):
        folderPath = os.path.join('')
    if (option & TEST_FILENAME == TEST_FILENAME):
        fileName = ""
    if (option & TEST_RANGE == TEST_RANGE):
        start = 4
        end = 2

    logging.info("---------------------- Program start ----------------------")
    logging.info("path : " + folderPath)
    logging.info("fileName : " + fileName + " / range : " + str(start) + " ~ " + str(end))
    examinationSTT(start, end, folderPath, fileName, "ETRI")

# -------------------- [Program Start] --------------------
if __name__ == '__main__':
    Run()

    # exceptionTest(TEST_RANGE)

    # 콘솔 메뉴 개발중 (UI 프로토타입)
    # showMenu()
    # select = -1
    # while(True):
    #     select = int(input("선택: "))
    #     if (select == 0):
    #         exit()

    #     if (select == 1):
    #         p_start = int(input("시작넘버: "))
    #         p_end = int(input("종료넘버: "))
    #         examinationSTT(p_start, p_end)